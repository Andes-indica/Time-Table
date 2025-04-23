
import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import csv
import glob
import os

# Constants for time management
TIME_INCREMENT = 30  # minutes
COURSE_PARAMETERS = {
    'LECTURE': 3,    # slots for lectures
    'LAB': 4,        # slots for labs
    'TUT': 2,        # slots for tutorials
    'SELF_STUDY': 2, # slots for self study
    'BUFFER': 1      # buffer between sessions
}

# Schedule Parameters - unchanged as they seem standard
CLASS_DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
DAY_START = time(9, 0)
DAY_END = time(18, 30)

# Meal break parameters
MEAL_PERIOD_START = time(12, 30)
MEAL_PERIOD_END = time(14, 0)
MEAL_DURATION = 60  # in minutes

# Initialize global variables
all_time_slots = []
meal_schedules = {}  # Dictionary to store meal times by semester

# Data loading functions
def try_load_csv(filename, encodings_list=None):
    """Generic CSV loader with robust error handling"""
    if encodings_list is None:
        encodings_list = ['utf-8-sig', 'utf-8', 'cp1252']
    
    data_frame = None
    error_message = None
    
    for encoding in encodings_list:
        try:
            data_frame = pd.read_csv(filename, encoding=encoding)
            # Clean data - replace empty and 'nan' strings
            data_frame = data_frame.replace(r'^\s*$', pd.NA, regex=True)
            data_frame = data_frame.replace('nan', pd.NA)
            return data_frame
        except UnicodeDecodeError:
            continue
        except Exception as e:
            error_message = str(e)
            continue
    
    if data_frame is None:
        print(f"Error: Failed to load {filename}.\nDetails: {error_message}")
        return pd.DataFrame()  # Return empty dataframe instead of exiting
        
    return data_frame

def import_facilities():
    """Load room information from CSV"""
    facilities = {}
    try:
        with open('rooms.csv', 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                facilities[row['id']] = {
                    'capacity': int(row['capacity']),
                    'type': row['type'],
                    'roomNumber': row['roomNumber'],
                    'schedule': {day_index: set() for day_index in range(len(CLASS_DAYS))}
                }
    except FileNotFoundError:
        print("Note: rooms.csv not found, using default room allocation")
        return None
    return facilities

def import_enrollment_data():
    """Load batch sizes and section information"""
    enrollment_data = {}
    
    # Regular class sizes
    try:
        df = pd.read_csv('updated_batches.csv')
        for _, row in df.iterrows():
            students_total = row['Total_Students']
            max_size = row['MaxBatchSize']
            
            # Calculate sections needed
            section_count = (students_total + max_size - 1) // max_size
            students_per_section = (students_total + section_count - 1) // section_count

            enrollment_data[(row['Department'], row['Semester'])] = {
                'total': students_total,
                'num_sections': section_count,
                'section_size': students_per_section
            }
    except FileNotFoundError:
        print("Note: updated_batches.csv not found, using default class sizes")
        
    # Elective course registrations
    try:
        elective_df = pd.read_csv('elective_registration.csv')
        for _, row in elective_df.iterrows():
            enrollment_data[('ELECTIVE', row['Course Code'])] = {
                'total': row['Total Students'],
                'num_sections': 1,  # Electives typically single section
                'section_size': row['Total Students']
            }
    except FileNotFoundError:
        print("Note: elective_registrations.csv not found")
        
    return enrollment_data

# Time management functions
def setup_time_slots():
    """Initialize the global time slots"""
    global all_time_slots
    all_time_slots = create_time_grid()

def create_time_grid():
    """Generate 30-minute time slots for the day"""
    time_grid = []
    current = datetime.combine(datetime.today(), DAY_START)
    end = datetime.combine(datetime.today(), DAY_END)
    
    while current < end:
        slot_start = current.time()
        slot_end = (current + timedelta(minutes=TIME_INCREMENT)).time()
        time_grid.append((slot_start, slot_end))
        current = current + timedelta(minutes=TIME_INCREMENT)
    
    return time_grid

def compute_meal_times(semesters):
    """Calculate staggered meal times for different semesters"""
    global meal_schedules
    meal_schedules = {}  # Reset global meal breaks
    
    if not semesters:
        return meal_schedules
        
    # Calculate spacing between meal periods
    total_minutes = (
        MEAL_PERIOD_END.hour * 60 + MEAL_PERIOD_END.minute -
        MEAL_PERIOD_START.hour * 60 - MEAL_PERIOD_START.minute
    )
    
    time_between = 0
    if len(semesters) > 1:
        time_between = (total_minutes - MEAL_DURATION) / (len(semesters) - 1)
    
    # Sort semesters for consistent assignment
    semesters_ordered = sorted(semesters)
    
    for idx, sem in enumerate(semesters_ordered):
        start_mins = (MEAL_PERIOD_START.hour * 60 + MEAL_PERIOD_START.minute + 
                      int(idx * time_between))
        
        start_hr = start_mins // 60
        start_min = start_mins % 60
        
        end_mins = start_mins + MEAL_DURATION
        end_hr = end_mins // 60
        end_min = end_mins % 60
        
        meal_schedules[sem] = (
            time(start_hr, start_min),
            time(end_hr, end_min)
        )
    
    return meal_schedules

def is_break_period(time_slot, semester=None):
    """Check if a time slot falls within designated break times"""
    global meal_schedules
    start_time, end_time = time_slot
    
    # Morning break: 10:30-11:00
    morning_break = (time(10, 30) <= start_time < time(11, 0))
    
    # Meal breaks based on semester
    is_meal_time = False
    if semester:
        base_sem = int(str(semester)[0])  # Extract first digit (e.g., 4 from 4A)
        if base_sem in meal_schedules:
            meal_start, meal_end = meal_schedules[base_sem]
            is_meal_time = (meal_start <= start_time < meal_end)
    else:
        # For general checks, block all meal periods
        is_meal_time = any(meal_start <= start_time < meal_end 
                        for meal_start, meal_end in meal_schedules.values())
    
    return morning_break or is_meal_time

# Course utility functions
def determine_required_sessions(course):
    """Calculate required sessions based on course credits and hours"""
    lecture_credits = float(course['L']) if pd.notna(course['L']) else 0
    tutorial_hours = int(course['T']) if pd.notna(course['T']) else 0
    lab_hours = int(course['P']) if pd.notna(course['P']) else 0
    self_study = int(course['S']) if pd.notna(course['S']) else 0
    
    # Check if self-study only course
    if self_study > 0 and lecture_credits == 0 and tutorial_hours == 0 and lab_hours == 0:
        return 0, 0, 0, 0
        
    # Calculate sessions based on credit hours
    lec_sessions = 0
    if lecture_credits > 0:
        lec_sessions = max(1, round(lecture_credits * 2/3))
    
    tut_sessions = tutorial_hours
    lab_sessions = lab_hours // 2
    ss_sessions = self_study // 4 if (lecture_credits > 0 or tutorial_hours > 0 or lab_hours > 0) else 0
    
    return lec_sessions, tut_sessions, lab_sessions, ss_sessions

def is_elective_course(code):
    """Check if a course is an elective based on code format"""
    return code.startswith('B') and '-' in code

def get_elective_group(code):
    """Extract elective group from course code"""
    if is_elective_course(code):
        return code.split('-')[0]
    return None

def find_group_slots(timetable, day, group):
    """Find existing slots with courses from the same elective group"""
    group_slots = []
    for slot_idx, slot in timetable[day].items():
        code = slot.get('code', '')
        if code and get_elective_group(code) == group:
            group_slots.append(slot_idx)
    return group_slots

def determine_course_priority(course):
    """Calculate scheduling priority based on course constraints"""
    priority_value = 0
    code = str(course['Course Code'])
    
    # Prioritize lab courses
    if pd.notna(course['P']) and course['P'] > 0 and not is_elective_course(code):
        priority_value += 10
        if 'CS' in code or 'EC' in code:  # CS/EC labs need special rooms
            priority_value += 2
    elif is_elective_course(code):
        priority_value += 1  # Lower priority for electives
    elif pd.notna(course['L']) and course['L'] > 2:
        priority_value += 3  # Regular lectures
    elif pd.notna(course['T']) and course['T'] > 0:
        priority_value += 2  # Tutorials
    return priority_value

def determine_room_type(course):
    """Determine required room type based on course needs"""
    if pd.notna(course['P']) and course['P'] > 0:
        course_code = str(course['Course Code']).upper()
        if 'CS' in course_code or 'DS' in course_code:
            return 'COMPUTER_LAB'
        elif 'EC' in course_code:
            return 'HARDWARE_LAB'
        return 'COMPUTER_LAB'  # Default for labs
    else:
        return 'LECTURE_ROOM'  # For lectures, tutorials, etc.

def choose_instructor(faculty_string):
    """Select a faculty from multiple possibilities"""
    if '/' in faculty_string:
        faculty_options = [f.strip() for f in faculty_string.split('/')]
        return faculty_options[0]  # Take first faculty as default
    return faculty_string

# Room allocation functions
def find_adjacent_room(room_id, rooms):
    """Find an adjacent lab room based on room numbering"""
    if not room_id:
        return None
    
    # Extract room number
    current_num = int(''.join(filter(str.isdigit, rooms[room_id]['roomNumber'])))
    current_floor = current_num // 100
    
    # Search for adjacent room with same type
    for rid, room in rooms.items():
        if rid != room_id and room['type'] == rooms[room_id]['type']:
            room_num = int(''.join(filter(str.isdigit, room['roomNumber'])))
            # Check if on same floor with adjacent number
            if room_num // 100 == current_floor and abs(room_num - current_num) == 1:
                return rid
    return None

def allocate_room(room_map, room_type, required_size, day, start_slot, duration, excluded_rooms):
    """Try to allocate a room of given type and size"""
    for room_id, room in room_map.items():
        if room_id in excluded_rooms or room['type'].upper() == 'LIBRARY':
            continue
            
        # Filter by room type
        if room_type in ['LEC', 'TUT', 'SELF_STUDY']:
            if not ('LECTURE_ROOM' in room['type'].upper() or 'SEATER' in room['type'].upper()):
                continue
        # For labs, match lab type exactly
        elif room_type == 'COMPUTER_LAB' and room['type'].upper() != 'COMPUTER_LAB':
            continue
        elif room_type == 'HARDWARE_LAB' and room['type'].upper() != 'HARDWARE_LAB':
            continue
            
        # Check capacity except for labs which can be split
        if room_type not in ['COMPUTER_LAB', 'HARDWARE_LAB'] and room['capacity'] < required_size:
            continue

        # Check availability
        is_available = True
        for i in range(duration):
            if start_slot + i in room['schedule'][day]:
                is_available = False
                break
                
        if is_available:
            for i in range(duration):
                room['schedule'][day].add(start_slot + i)
            return room_id
                
    return None

def assign_suitable_room(course_type, department, semester, day, start_slot, duration, 
                       rooms, enrollment_data, timetable, course_code="", excluded_rooms=None):
    """Find suitable room(s) considering student counts and constraints"""
    if not rooms:
        return "DEFAULT_ROOM"
    
    required_size = 60  # Default fallback
    is_elective = is_elective_course(course_code)
    
    if enrollment_data:
        # For electives, check enrollment data
        if is_elective:
            elective_info = enrollment_data.get(('ELECTIVE', course_code))
            if elective_info:
                required_size = elective_info['section_size']
        else:
            # For regular courses use department enrollment data
            dept_info = enrollment_data.get((department, semester))
            if dept_info:
                required_size = dept_info['section_size']

    rooms_to_exclude = set() if excluded_rooms is None else excluded_rooms

    # Special handling for labs that may need adjacent rooms
    if course_type in ['COMPUTER_LAB', 'HARDWARE_LAB']:
        dept_info = enrollment_data.get((department, semester))
        if dept_info and dept_info['total'] > 35:  # Standard lab capacity
            # Try to find adjacent lab rooms
            for room_id, room in rooms.items():
                if room_id in rooms_to_exclude or room['type'].upper() != course_type:
                    continue
                    
                # Check if this room is available
                is_available = True
                for i in range(duration):
                    if start_slot + i in room['schedule'][day]:
                        is_available = False
                        break
                
                if is_available:
                    # Try to find an adjacent room
                    adjacent_room = find_adjacent_room(room_id, rooms)
                    if adjacent_room and adjacent_room not in rooms_to_exclude:
                        # Check if adjacent room is also available
                        adjacent_available = True
                        for i in range(duration):
                            if start_slot + i in rooms[adjacent_room]['schedule'][day]:
                                adjacent_available = False
                                break
                        
                        if adjacent_available:
                            # Mark both rooms as used
                            for i in range(duration):
                                room['schedule'][day].add(start_slot + i)
                                rooms[adjacent_room]['schedule'][day].add(start_slot + i)
                            return f"{room_id},{adjacent_room}"  # Return both room IDs
                            
        # If we don't need two rooms or couldn't find adjacent ones, use regular allocation
        return allocate_room(rooms, course_type, required_size, day, start_slot, duration, rooms_to_exclude)

    # For lectures and elective courses
    if course_type in ['LEC', 'TUT', 'SELF_STUDY'] or is_elective:
        # First try regular lecture rooms
        lecture_rooms = {rid: room for rid, room in rooms.items() 
                        if 'LECTURE_ROOM' in room['type'].upper()}
        
        # Then try large seater rooms 
        seater_rooms = {rid: room for rid, room in rooms.items()
                       if 'SEATER' in room['type'].upper()}
        
        # Special handling for elective courses
        if is_elective:
            elective_group = get_elective_group(course_code)
            elective_excluded_rooms = set()
            elective_group_rooms = {}  # Track rooms already used by this elective group
            
            # Track room usage count
            room_usage = {rid: sum(len(room['schedule'][d]) for d in range(len(CLASS_DAYS))) 
                         for rid, room in rooms.items()}
            
            # Sort rooms by usage count
            sorted_lecture_rooms = dict(sorted(lecture_rooms.items(), 
                                             key=lambda x: room_usage[x[0]]))
            sorted_seater_rooms = dict(sorted(seater_rooms.items(),
                                            key=lambda x: room_usage[x[0]]))
            
            # Check availability for the sorted rooms
            for room_dict in [sorted_lecture_rooms, sorted_seater_rooms]:
                for room_id, room in room_dict.items():
                    is_occupied = False
                    for slot in range(start_slot, start_slot + duration):
                        if slot in rooms[room_id]['schedule'][day]:
                            # Check if room is used by any course from same elective group
                            if slot in timetable[day]:
                                slot_data = timetable[day][slot]
                                if (slot_data['classroom'] == room_id and 
                                    slot_data['type'] is not None):
                                    slot_code = slot_data.get('code', '')
                                    if get_elective_group(slot_code) == elective_group:
                                        elective_group_rooms[slot_code] = room_id
                                    else:
                                        elective_excluded_rooms.add(room_id)
                            is_occupied = True
                            break
                    
                    # Room is free for this time slot
                    if not is_occupied and room_id not in elective_excluded_rooms:
                        if 'capacity' in room and room['capacity'] >= required_size:
                            # Mark slots as used
                            for i in range(duration):
                                room['schedule'][day].add(start_slot + i)
                            return room_id
            
            # If no unused room found, try existing elective group rooms
            if course_code in elective_group_rooms:
                return elective_group_rooms[course_code]
            
            # Try remaining rooms through regular allocation
            room_id = allocate_room(lecture_rooms, 'LEC', required_size,
                                  day, start_slot, duration, elective_excluded_rooms)
            
            if not room_id:
                room_id = allocate_room(seater_rooms, 'LEC', required_size,
                                      day, start_slot, duration, elective_excluded_rooms)
            
            if room_id:
                elective_group_rooms[course_code] = room_id
            
            return room_id

        # For regular courses, use standard logic
        room_id = allocate_room(lecture_rooms, 'LEC', required_size,
                              day, start_slot, duration, rooms_to_exclude)
        if not room_id:
            room_id = allocate_room(seater_rooms, 'LEC', required_size,
                                  day, start_slot, duration, rooms_to_exclude)
        return room_id
    
    # For labs, use standard allocation logic
    return allocate_room(rooms, course_type, required_size,
                       day, start_slot, duration, rooms_to_exclude)

# Scheduling constraint checking
def is_activity_scheduled(timetable, day, start_slot, end_slot):
    """Check if there's an activity scheduled in the given time range"""
    for slot in range(start_slot, end_slot):
        if (slot < len(timetable[day]) and 
            timetable[day][slot]['type'] and 
            timetable[day][slot]['type'] in ['LEC', 'LAB', 'TUT']):
            return True
    return False

def check_instructor_workload(instructor_schedule, instructor, day, department, semester, 
                           section, timetable, course_code=None, activity_type=None):
    """Check instructor scheduling constraints for the day"""
    sessions_count = 0
    instructor_courses = set()  # Track instructor's courses
    
    # Count all sessions for this instructor on this day
    for slot in timetable[day].values():
        if slot['faculty'] == instructor and slot['type'] in ['LEC', 'LAB', 'TUT']:
            slot_code = slot.get('code', '')
            if slot_code:
                # For regular courses
                if not is_elective_course(slot_code):
                    sessions_count += 1
                # For electives, only count each course once
                elif slot_code not in instructor_courses:
                    sessions_count += 1
                    instructor_courses.add(slot_code)
                    
    # Special handling for electives - allow parallel scheduling
    if course_code and is_elective_course(course_code):
        elective_group = get_elective_group(course_code)
        existing_slots = find_group_slots(timetable, day, elective_group)
        if existing_slots:
            # For electives, allow more flexibility
            return sessions_count < 3
    
    return sessions_count < 2  # Standard limit for regular courses

def check_course_session_spacing(instructor_schedule, timetable, instructor, course_code, day, start_slot):
    """Check if there is sufficient gap between sessions of the same course"""
    min_hours_gap = 3
    slots_per_hour = 2  # 30-min slots
    gap_slots = min_hours_gap * slots_per_hour
    
    # Check previous slots
    for i in range(max(0, start_slot - gap_slots), start_slot):
        if i in instructor_schedule[instructor][day]:
            slot_data = timetable[day][i]
            if slot_data['code'] == course_code and slot_data['type'] in ['LEC', 'TUT']:
                return False
                
    # Check upcoming slots  
    for i in range(start_slot + 1, min(len(all_time_slots), start_slot + gap_slots)):
        if i in instructor_schedule[instructor][day]:
            slot_data = timetable[day][i]
            if slot_data['code'] == course_code and slot_data['type'] in ['LEC', 'TUT']:
                return False
    
    return True

def find_available_slots(timetable, instructor_schedule, instructor, day, duration, 
                       reserved_slots, semester, department):
    """Find available consecutive slots in a day"""
    available_slots = []
    
    for start_slot in range(len(all_time_slots) - duration + 1):
        is_available = True
        # Check each slot in the duration
        for i in range(duration):
            current_slot = start_slot + i
            if (current_slot in instructor_schedule[instructor][day] or
                timetable[day][current_slot]['type'] is not None or
                is_break_period(all_time_slots[current_slot], semester) or
                is_slot_reserved(all_time_slots[current_slot], CLASS_DAYS[day], 
                               semester, department, reserved_slots)):
                is_available = False
                break

        if is_available:
            available_slots.append(start_slot)
    
    return available_slots

def is_slot_reserved(slot, day, semester, department, reserved_slots):
    """Check if a time slot is reserved - placeholder function"""
    # This function always returns False since we're not using reserved_slots.csv
    return False

def load_reserved_slots():
    """Return empty reserved slots structure"""
    return {day: {} for day in CLASS_DAYS}

# Output functions
def display_unscheduled_summary(unscheduled_list):
    """Print a summary of unscheduled courses to the console"""
    if not unscheduled_list:
        print("All courses were successfully scheduled!")
        return
        
    print("\n" + "="*80)
    print("UNSCHEDULED COURSES SUMMARY".center(80))
    print("="*80)
    print(f"{'Department':<10} {'Semester':<10} {'Course Code':<15} {'Course Name':<30} {'Faculty':<20} {'Missing':<8}")
    print("-"*80)
    
    for course in unscheduled_list:
        missing = course['Expected Slots'] - course['Scheduled Slots']
        print(f"{course['Department']:<10} {course['Semester']:<10} {course['Code']:<15} {course['Name'][:28]:<30} {course['Faculty'][:18]:<20} {missing:<8}")
    
    print("="*80)
    print(f"Total unscheduled courses: {len(unscheduled_list)}")
    print("="*80)

# Main timetable generation function moved to top level
def generate_all_timetables():
    global meal_schedules
    setup_time_slots()
    reserved_slots = load_reserved_slots()
    workbook = Workbook()
    workbook.remove(workbook.active)
    instructor_schedules = {}
    facilities = import_facilities()
    enrollment_data = import_enrollment_data()

    # Track unscheduled courses
    courses_not_scheduled = []

    # Get main course data with robust error handling
    course_data = try_load_csv('combined.csv')
    if course_data.empty:
        print("Error: No valid data found in combined.csv")
        return ["error.xlsx"]  # Return a dummy filename

    # Get all unique semester numbers
    all_semester_bases = sorted(set(int(str(sem)[0]) for sem in course_data['Semester'].unique()))
    # Calculate meal breaks dynamically
    meal_schedules = compute_meal_times(all_semester_bases)

    for department in course_data['Department'].unique():
        # Process all semesters for this department
        for semester in course_data[course_data['Department'] == department]['Semester'].unique():
            # Filter out courses marked as not to be scheduled
            active_courses = course_data[(course_data['Department'] == department) & 
                           (course_data['Semester'] == semester) & 
                           ((course_data['Schedule'].fillna('Yes').str.upper() == 'YES') | 
                            (course_data['Schedule'].isna()))].copy()
            
            if active_courses.empty:
                continue

            # First process lab courses (higher priority)
            lab_courses = active_courses[active_courses['P'] > 0].copy()
            lab_courses['priority'] = lab_courses.apply(determine_course_priority, axis=1)
            lab_courses = lab_courses.sort_values('priority', ascending=False)

            # Then process non-lab courses
            regular_courses = active_courses[active_courses['P'] == 0].copy()
            regular_courses['priority'] = regular_courses.apply(determine_course_priority, axis=1)
            regular_courses = regular_courses.sort_values('priority', ascending=False)

            # Combine for processing with labs first
            prioritized_courses = pd.concat([lab_courses, regular_courses])

            # Get section info
            dept_enrollment = enrollment_data.get((department, semester))
            section_count = dept_enrollment['num_sections'] if dept_enrollment else 1

            for section_idx in range(section_count):
                # Create sheet name based on sections
                if section_count == 1:
                    sheet_name = f"{department}{semester}"
                else:
                    sheet_name = f"{department}{semester}_{chr(65+section_idx)}"
                
                worksheet = workbook.create_sheet(title=sheet_name)
                
                # Initialize timetable structure
                schedule = {day_idx: {slot_idx: {
                            'type': None, 
                            'code': '', 
                            'name': '', 
                            'faculty': '', 
                            'classroom': ''
                         } for slot_idx in range(len(all_time_slots))
                       } for day_idx in range(len(CLASS_DAYS))}
                
                # Prioritize courses
                prioritized_courses['priority'] = prioritized_courses.apply(determine_course_priority, axis=1)
                prioritized_courses = prioritized_courses.sort_values('priority', ascending=False)

                # Schedule all courses
                for _, course in prioritized_courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    instructor = str(course['Faculty'])
                    
                    # Calculate required sessions
                    lec_sessions, tut_sessions, lab_sessions, self_study = determine_required_sessions(course)
                    
                    if instructor not in instructor_schedules:
                        instructor_schedules[instructor] = {day_idx: set() for day_idx in range(len(CLASS_DAYS))}

                    # Process lecture sessions
                    for _ in range(lec_sessions):
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < 1000:
                            day_idx = random.randint(0, len(CLASS_DAYS)-1)
                            start_slot = random.randint(0, len(all_time_slots)-COURSE_PARAMETERS['LECTURE'])
                            
                            # Check for adequate spacing between course sessions
                            if not check_course_session_spacing(instructor_schedules, schedule, 
                                                              instructor, code, day_idx, start_slot):
                                attempts += 1
                                continue
                            
                            # Check if any slot is reserved
                            any_reserved = any(is_slot_reserved(all_time_slots[start_slot + i], 
                                                              CLASS_DAYS[day_idx],
                                                              semester,
                                                              department,
                                                              reserved_slots) 
                                             for i in range(COURSE_PARAMETERS['LECTURE']))
                            
                            if any_reserved:
                                attempts += 1
                                continue
                            
                            # Check instructor workload limits
                            if not check_instructor_workload(instructor_schedules, instructor, day_idx, 
                                                          department, semester, section_idx, schedule,
                                                          code, 'LEC'):
                                attempts += 1
                                continue
                                
                            # Check availability and ensure breaks between sessions
                            slots_available = True
                            for i in range(COURSE_PARAMETERS['LECTURE']):
                                slot_idx = start_slot + i
                                if (slot_idx in instructor_schedules[instructor][day_idx] or 
                                    schedule[day_idx][slot_idx]['type'] is not None or
                                    is_break_period(all_time_slots[slot_idx], semester)):
                                    slots_available = False
                                    break
                                
                                # Check for sessions adjacent to this slot
                                if slot_idx > 0:
                                    if is_activity_scheduled(schedule, day_idx, 
                                                           max(0, slot_idx - COURSE_PARAMETERS['BUFFER']), 
                                                           slot_idx):
                                        slots_available = False
                                        break
                                
                                if slot_idx < len(all_time_slots) - 1:
                                    if is_activity_scheduled(schedule, day_idx,
                                                           slot_idx + 1,
                                                           min(len(all_time_slots), 
                                                               slot_idx + COURSE_PARAMETERS['BUFFER'] + 1)):
                                        slots_available = False
                                        break
                            
                            if slots_available:
                                room_id = assign_suitable_room('LEC', department, semester, 
                                                            day_idx, start_slot, COURSE_PARAMETERS['LECTURE'], 
                                                            facilities, enrollment_data, schedule, code)
                                
                                if room_id:
                                    # Mark slots as used
                                    for i in range(COURSE_PARAMETERS['LECTURE']):
                                        instructor_schedules[instructor][day_idx].add(start_slot+i)
                                        schedule[day_idx][start_slot+i]['type'] = 'LEC'
                                        schedule[day_idx][start_slot+i]['code'] = code if i == 0 else ''
                                        schedule[day_idx][start_slot+i]['name'] = name if i == 0 else ''
                                        schedule[day_idx][start_slot+i]['faculty'] = instructor if i == 0 else ''
                                        schedule[day_idx][start_slot+i]['classroom'] = room_id if i == 0 else ''
                                    scheduled = True
                            attempts += 1

                    # Process tutorial sessions
                    for _ in range(tut_sessions):
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < 1000:
                            day_idx = random.randint(0, len(CLASS_DAYS)-1)
                            
                            # Check spacing between sessions
                            if not check_course_session_spacing(instructor_schedules, schedule, 
                                                             instructor, code, day_idx, start_slot):
                                attempts += 1
                                continue
                            
                            # Check instructor workload
                            if not check_instructor_workload(instructor_schedules, instructor, day_idx,
                                                          department, semester, section_idx, schedule,
                                                          code, 'TUT'):
                                attempts += 1
                                continue
                                
                            start_slot = random.randint(0, len(all_time_slots)-COURSE_PARAMETERS['TUT'])
                            
                            # Check for reserved slots
                            any_reserved = any(is_slot_reserved(all_time_slots[start_slot + i], 
                                                              CLASS_DAYS[day_idx],
                                                              semester,
                                                              department,
                                                              reserved_slots) 
                                             for i in range(COURSE_PARAMETERS['TUT']))
                            
                            if any_reserved:
                                attempts += 1
                                continue
                            
                            # Check availability
                            slots_available = True
                            for i in range(COURSE_PARAMETERS['TUT']):
                                if (start_slot+i in instructor_schedules[instructor][day_idx] or 
                                    schedule[day_idx][start_slot+i]['type'] is not None or
                                    is_break_period(all_time_slots[start_slot+i], semester)):
                                    slots_available = False
                                    break
                            
                            if slots_available:
                                room_id = assign_suitable_room('TUT', department, semester, 
                                                            day_idx, start_slot, COURSE_PARAMETERS['TUT'], 
                                                            facilities, enrollment_data, schedule, code)
                                
                                if room_id:
                                    # Mark slots as used
                                    for i in range(COURSE_PARAMETERS['TUT']):
                                        instructor_schedules[instructor][day_idx].add(start_slot+i)
                                        schedule[day_idx][start_slot+i]['type'] = 'TUT'
                                        schedule[day_idx][start_slot+i]['code'] = code if i == 0 else ''
                                        schedule[day_idx][start_slot+i]['name'] = name if i == 0 else ''
                                        schedule[day_idx][start_slot+i]['faculty'] = instructor if i == 0 else ''
                                        schedule[day_idx][start_slot+i]['classroom'] = room_id if i == 0 else ''
                                    scheduled = True
                            attempts += 1

                    # Process lab sessions
                    if lab_sessions > 0:
                        room_type = determine_room_type(course)
                        for _ in range(lab_sessions):
                            scheduled = False
                            attempts = 0
                            
                            # Try days in random order
                            day_options = list(range(len(CLASS_DAYS)))
                            random.shuffle(day_options)
                            
                            for day_idx in day_options:
                                # Get available slots for this day
                                available_slots = find_available_slots(
                                    schedule, instructor_schedules, instructor, day_idx, 
                                    COURSE_PARAMETERS['LAB'], reserved_slots, semester, department
                                )
                                
                                for start_slot in available_slots:
                                    room_id = assign_suitable_room(
                                        room_type, department, semester, day_idx, start_slot, 
                                        COURSE_PARAMETERS['LAB'], facilities, enrollment_data, 
                                        schedule, code
                                    )
                                    
                                    if room_id:
                                        # Format room display for paired labs
                                        display_room = room_id
                                        if ',' in str(room_id):
                                            room1, room2 = room_id.split(',')
                                            display_room = f"{room1}+{room2}"
                                        
                                        # Mark slots as used
                                        for i in range(COURSE_PARAMETERS['LAB']):
                                            instructor_schedules[instructor][day_idx].add(start_slot+i)
                                            schedule[day_idx][start_slot+i]['type'] = 'LAB'
                                            schedule[day_idx][start_slot+i]['code'] = code if i == 0 else ''
                                            schedule[day_idx][start_slot+i]['name'] = name if i == 0 else ''
                                            schedule[day_idx][start_slot+i]['faculty'] = instructor if i == 0 else ''
                                            schedule[day_idx][start_slot+i]['classroom'] = display_room if i == 0 else ''
                                        scheduled = True
                                        break
                                
                                if scheduled:
                                    break

                # Process self-study sessions
                for _, course in prioritized_courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    instructor = str(course['Faculty'])
                    _, _, _, self_study = determine_required_sessions(course)
                    
                    if self_study > 0:
                        if instructor not in instructor_schedules:
                            instructor_schedules[instructor] = {day_idx: set() for day_idx in range(len(CLASS_DAYS))}
                        
                        # Schedule each self-study session
                        for _ in range(self_study):
                            scheduled = False
                            attempts = 0
                            while not scheduled and attempts < 1000:
                                day_idx = random.randint(0, len(CLASS_DAYS)-1)
                                start_slot = random.randint(0, len(all_time_slots)-COURSE_PARAMETERS['SELF_STUDY'])
                                
                                # Check for reserved slots
                                any_reserved = any(is_slot_reserved(all_time_slots[start_slot + i], 
                                                                  CLASS_DAYS[day_idx],
                                                                  semester,
                                                                  department,
                                                                  reserved_slots) 
                                                 for i in range(COURSE_PARAMETERS['SELF_STUDY']))
                                
                                if any_reserved:
                                    attempts += 1
                                    continue
                                
                                # Check availability
                                slots_available = True
                                for i in range(COURSE_PARAMETERS['SELF_STUDY']):
                                    if (start_slot+i in instructor_schedules[instructor][day_idx] or 
                                        schedule[day_idx][start_slot+i]['type'] is not None or
                                        is_break_period(all_time_slots[start_slot+i], semester)):
                                        slots_available = False
                                        break
                                
                                if slots_available:
                                    room_id = assign_suitable_room('SELF_STUDY', department, semester, 
                                                                day_idx, start_slot, COURSE_PARAMETERS['SELF_STUDY'], 
                                                                facilities, enrollment_data, schedule, code)
                                    
                                    if room_id:
                                        # Mark slots as used
                                        for i in range(COURSE_PARAMETERS['SELF_STUDY']):
                                            instructor_schedules[instructor][day_idx].add(start_slot+i)
                                            schedule[day_idx][start_slot+i]['type'] = 'SS'  # Self Study
                                            schedule[day_idx][start_slot+i]['code'] = code if i == 0 else ''
                                            schedule[day_idx][start_slot+i]['name'] = name if i == 0 else ''
                                            schedule[day_idx][start_slot+i]['faculty'] = instructor if i == 0 else ''
                                            schedule[day_idx][start_slot+i]['classroom'] = room_id if i == 0 else ''
                                        scheduled = True
                                attempts += 1

                # Write timetable to worksheet - applying styles and formatting
                time_labels = ['Day'] + [f"{t[0].strftime('%H:%M')}-{t[1].strftime('%H:%M')}" for t in all_time_slots]
                worksheet.append(time_labels)
                
                # Style header row
                header_font = Font(bold=True)
                center_align = Alignment(horizontal='center', vertical='center')
                
                for header_cell in worksheet[1]:
                    header_cell.font = header_font
                    header_cell.alignment = center_align
                
                # Activity color scheme
                style_lecture = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
                style_lab = PatternFill(start_color="FAE5D3", end_color="FAE5D3", fill_type="solid")
                style_tutorial = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
                
                cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                
                # Generate timetable grid
                for day_idx, day_name in enumerate(CLASS_DAYS):
                    row_idx = day_idx + 2
                    worksheet.append([day_name])
                    
                    merge_cells = []  # Track cells to merge
                    
                    for slot_idx in range(len(all_time_slots)):
                        content = ''
                        cell_style = None
                        
                        if is_break_period(all_time_slots[slot_idx], semester):
                            content = "BREAK"
                        elif schedule[day_idx][slot_idx]['type']:
                            activity = schedule[day_idx][slot_idx]['type']
                            course_code = schedule[day_idx][slot_idx]['code']
                            room = schedule[day_idx][slot_idx]['classroom']
                            faculty = schedule[day_idx][slot_idx]['faculty']
                            
                            if course_code:
                                session_length = {
                                    'LEC': COURSE_PARAMETERS['LECTURE'],
                                    'LAB': COURSE_PARAMETERS['LAB'],
                                    'TUT': COURSE_PARAMETERS['TUT'],
                                    'SS': COURSE_PARAMETERS['SELF_STUDY']
                                }.get(activity, 1)
                                
                                # Apply appropriate style based on activity
                                cell_style = {
                                    'LEC': style_lecture,
                                    'LAB': style_lab,
                                    'TUT': style_tutorial
                                }.get(activity)
                                
                                content = f"{course_code} {activity}\n{room}\n{faculty}"
                                
                                # Create merge range if activity spans multiple slots
                                if session_length > 1:
                                    start_col = get_column_letter(slot_idx + 2)
                                    end_col = get_column_letter(slot_idx + session_length + 1)
                                    merge_range = f"{start_col}{row_idx}:{end_col}{row_idx}"
                                    merge_cells.append((merge_range, cell_style))
                        
                        cell = worksheet.cell(row=row_idx, column=slot_idx+2, value=content)
                        if cell_style:
                            cell.fill = cell_style
                        cell.border = cell_border
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    
                    # Apply cell merges after creating all cells in the row
                    for merge_range, style in merge_cells:
                        worksheet.merge_cells(merge_range)
                        merged_cell = worksheet[merge_range.split(':')[0]]
                        if style:
                            merged_cell.fill = style
                        merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

                # Set column widths and row heights
                for col_idx in range(1, len(all_time_slots)+2):
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = 15
                
                for row in worksheet.iter_rows(min_row=2, max_row=len(CLASS_DAYS)+1):
                    worksheet.row_dimensions[row[0].row].height = 40

                # Add unscheduled courses section
                dept_courses = course_data[(course_data['Department'] == department) & 
                                       (course_data['Semester'] == semester) &
                                       ((course_data['Schedule'].fillna('Yes').str.upper() == 'YES') | 
                                        (course_data['Schedule'].isna()))].copy()

                # Add spacing
                worksheet.append([])
                worksheet.append([])

                # Add unscheduled courses header
                header_row = worksheet.max_row + 1
                worksheet.cell(row=header_row, column=1, value="Unscheduled Courses").font = Font(bold=True)
                worksheet.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=6)

                # Add column headers
                header_labels = ['Course Code', 'Course Name', 'Faculty', 'Required Components', 'Missing Components']
                worksheet.append(header_labels)
                for idx, header in enumerate(header_labels, 1):
                    cell = worksheet.cell(row=header_row + 1, column=idx)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # Track unscheduled components
                for _, course in dept_courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    faculty = str(course['Faculty'])
                    
                    # Calculate required components
                    lec_sessions, tut_sessions, lab_sessions, self_study = determine_required_sessions(course)
                    required_components = []
                    if lec_sessions > 0: required_components.append(f"LEC:{lec_sessions}")
                    if tut_sessions > 0: required_components.append(f"TUT:{tut_sessions}")
                    if lab_sessions > 0: required_components.append(f"LAB:{lab_sessions}")
                    if self_study > 0: required_components.append(f"SS:{self_study}")
                    
                    # Count scheduled components
                    scheduled_lec = sum(1 for d in range(len(CLASS_DAYS)) 
                                     for s in range(len(all_time_slots)) 
                                     if schedule[d][s]['code'] == code 
                                     and schedule[d][s]['type'] == 'LEC')
                    
                    scheduled_tut = sum(1 for d in range(len(CLASS_DAYS))
                                     for s in range(len(all_time_slots))
                                     if schedule[d][s]['code'] == code
                                     and schedule[d][s]['type'] == 'TUT')
                    
                    scheduled_lab = sum(1 for d in range(len(CLASS_DAYS))
                                     for s in range(len(all_time_slots))
                                     if schedule[d][s]['code'] == code
                                     and schedule[d][s]['type'] == 'LAB')
                    
                    scheduled_ss = sum(1 for d in range(len(CLASS_DAYS))
                                    for s in range(len(all_time_slots))
                                    if schedule[d][s]['code'] == code
                                    and schedule[d][s]['type'] == 'SS')
                    
                    # Calculate missing components
                    missing_components = []
                    if scheduled_lec < lec_sessions: missing_components.append(f"LEC:{lec_sessions-scheduled_lec}")
                    if scheduled_tut < tut_sessions: missing_components.append(f"TUT:{tut_sessions-scheduled_tut}")
                    if scheduled_lab < lab_sessions: missing_components.append(f"LAB:{lab_sessions-scheduled_lab}")
                    if scheduled_ss < self_study: missing_components.append(f"SS:{self_study-scheduled_ss}")
                    
                    # Add row if there are missing components
                    if missing_components:
                        worksheet.append([
                            code,
                            name,
                            faculty,
                            ', '.join(required_components),
                            ', '.join(missing_components)
                        ])
                        
                        # Add to global unscheduled list
                        courses_not_scheduled.append({
                            'Department': department,
                            'Semester': semester,
                            'Code': code,
                            'Name': name,
                            'Faculty': faculty,
                            'Expected Slots': lec_sessions + tut_sessions + lab_sessions + self_study,
                            'Scheduled Slots': scheduled_lec + scheduled_tut + scheduled_lab + scheduled_ss
                        })

                # Style the unscheduled courses section
                for row in worksheet.iter_rows(min_row=header_row, max_row=worksheet.max_row):
                    for cell in row:
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                           top=Side(style='thin'), bottom=Side(style='thin'))
                        cell.alignment = Alignment(horizontal='center')

                # Adjust column widths for the unscheduled section
                for col in range(1, 6):
                    worksheet.column_dimensions[get_column_letter(col)].width = 20

    # Create unscheduled courses summary sheet
    if courses_not_scheduled:
        summary_sheet = workbook.create_sheet(title="Unscheduled Summary")
        
        # Add headers
        summary_headers = ['Department', 'Semester', 'Course Code', 'Course Name', 
                         'Faculty', 'Expected Slots', 'Scheduled Slots', 'Missing Slots']
        summary_sheet.append(summary_headers)
        
        # Style headers
        for cell in summary_sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for course in courses_not_scheduled:
            summary_sheet.append([
                course['Department'],
                course['Semester'],
                course['Code'],
                course['Name'],
                course['Faculty'],
                course['Expected Slots'],
                course['Scheduled Slots'],
                course['Expected Slots'] - course['Scheduled Slots']
            ])
        
        # Style data cells
        for row in summary_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in summary_sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            summary_sheet.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2

    # Save workbook with error handling
    output_filename = "timetable_all.xlsx"
    try:
        # Try to save the file
        workbook.save(output_filename)
        print(f"Complete timetable saved as {output_filename}")
    except PermissionError:
        # If file is open/locked, try saving with a new name
        base, ext = os.path.splitext(output_filename)
        counter = 1
        while True:
            new_filename = f"{base}_{counter}{ext}"
            try:
                workbook.save(new_filename)
                print(f"File was locked. Saved as {new_filename} instead")
                output_filename = new_filename
                break
            except PermissionError:
                counter += 1
                if counter > 100:  # Prevent infinite loop
                    raise Exception("Unable to save file after 100 attempts")
    
    # Print unscheduled courses summary to console
    display_unscheduled_summary(courses_not_scheduled)
    
    return [output_filename]

if __name__ == "__main__":
    generate_all_timetables()